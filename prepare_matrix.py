import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import numpy as np

prepo = pd.read_csv('pre-processed.csv')
noi = pd.read_csv('number_of_isolates.csv')

with pd.ExcelWriter("amb.xlsx", engine="openpyxl") as writer:
    for specimen_type in prepo['specimen_category'].unique():
        df = prepo[prepo['specimen_category']==specimen_type]
        print(specimen_type)
        #make it a matrix
        amb_matrix_perc = pd.pivot(df, index='pathogen', columns='antibiotic', values='sensitivity')
        amb_matrix_tot = pd.pivot(df, index='pathogen', columns='antibiotic', values='total')

        full_amb_list = []
        total_pat_entry = dict()
        total_n_isolates = 0

        for pat in amb_matrix_perc.index:
            this_pat_entry = dict()
            weigh_sum = 0 #total number of *tests results* for all antibiotic. It is not number of isolates since one isolate is tested on multiple antibiotics
            val_avg = 0 #numerator of weighted average of sensitivity
            for antibio in amb_matrix_perc.loc[pat].index:
                if np.isnan(amb_matrix_tot.loc[pat][antibio]):
                    this_pat_entry[(antibio,'n')]=0
                    this_pat_entry[(antibio,'s')]=np.nan
                else:
                    no_iso = amb_matrix_tot.loc[pat][antibio]
                    this_pat_entry[(antibio,'n')]=no_iso
                    this_pat_entry[(antibio,'s')]=str(round(100*amb_matrix_perc.loc[pat][antibio]))+'%'
                    if (antibio,'n') in total_pat_entry:
                        total_pat_entry[(antibio,'n')]=total_pat_entry[(antibio,'n')]+ no_iso
                        total_pat_entry[(antibio,'s')]=total_pat_entry[(antibio,'s')] + no_iso * amb_matrix_perc.loc[pat][antibio]
                    else:
                        total_pat_entry[(antibio,'n')]=no_iso
                        total_pat_entry[(antibio,'s')]=no_iso * amb_matrix_perc.loc[pat][antibio]

                    weigh_sum = weigh_sum + no_iso
                    val_avg = val_avg + no_iso * amb_matrix_perc.loc[pat][antibio]

            this_pat_entry[('total', 'weigh_avg_sens')]=str(round(100 * val_avg/weigh_sum)) + '%'
            n_of_isolates = noi[(noi['specimen_category']==specimen_type) & (noi['pathogen']==pat)]['number_of_isolates'].iloc[0]
            this_pat_entry[('total', 'n_isolates')]=n_of_isolates
            total_n_isolates = total_n_isolates + n_of_isolates
            full_amb_list.append(this_pat_entry)

        for antibio in amb_matrix_perc.loc[pat].index:
            total_pat_entry[(antibio,'s')]=str(round(100*(total_pat_entry[(antibio,'s')]/total_pat_entry[(antibio,'n')]))) + '%'
        total_pat_entry[('total','n_isolates')]=total_n_isolates
        full_amb_list.append(total_pat_entry)



        full_amb_index = list(amb_matrix_perc.index)
        full_amb_index.append('Total') #append is in-place

        full_amb = pd.DataFrame.from_records(full_amb_list,index=full_amb_index)
        new_columns = pd.MultiIndex.from_tuples(full_amb.columns, names=['Antibiotic', 'number of isolates and sensitivity'])
        full_amb.columns = new_columns
        full_amb.fillna('--', inplace=True)
        full_amb.to_excel(writer,sheet_name=specimen_type)
        # Set backgrund colors depending on cell values
        heading_size = 3
        worksheet = writer.sheets[specimen_type]

        worksheet.row_dimensions[1].height = 150

        worksheet.column_dimensions[worksheet.cell(row=3,column=1).column_letter].width = 40
        for ccc in range(2,len(full_amb.columns)+1):
            if ccc%2:
                worksheet.column_dimensions[worksheet.cell(row=3,column=ccc).column_letter].width = 5
            else:
                worksheet.column_dimensions[worksheet.cell(row=3,column=ccc).column_letter].width = 5
            worksheet.cell(row=1,column=ccc).alignment = Alignment(textRotation=70)
            for rrr in range(heading_size+1,heading_size+1+len(full_amb)):
                worksheet.cell(row=rrr,column=ccc).alignment = Alignment(vertical='center')
                worksheet.cell(row=rrr,column=ccc).alignment = Alignment(horizontal='center')

        worksheet.column_dimensions[worksheet.cell(row=3,column=len(full_amb.columns)+1).column_letter].width = 20
        worksheet.column_dimensions[worksheet.cell(row=3,column=len(full_amb.columns)).column_letter].width = 20

        for rrr in range(heading_size+1,heading_size+1+len(full_amb)):
            for ccc in range(3,len(full_amb.columns)+1,2):
                value_str = worksheet.cell(row=rrr,column=ccc).value
                if value_str == '--':
                    worksheet.cell(row=rrr,column=ccc).fill = PatternFill("solid", start_color=('ededed'))
                else:
                    value = int(value_str.rstrip('%'))
                    worksheet.cell(row=rrr,column=ccc).fill = PatternFill("solid", start_color=('d62d20' if value < 50 else 'd7503c' if value < 75 else 'eee600' if value < 90 else 'a0e040'))
        worksheet.delete_rows(3,1)
